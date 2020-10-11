import requests as re
from bs4 import BeautifulSoup as bs
import openpyxl
import unicodedata as ud
import os

url_prod = "https://www.e-katalog.ru/GIGABYTE-X299X-AORUS-XTREME-WATERFORCE.htm"


# url_prod ="https://www.e-katalog.ru/SAMSUNG-KVADRAT-COVER-FOR-GALAXY-NOTE20-ULTRA.htm"
# url_prod = "https://www.e-katalog.ru/PLANTRONICS-EXPLORER-ML15.htm"
# url_prod = "https://www.e-katalog.ru/ATLANT-XM-4208-000.htm"

def product(url_prod):
    prod = re.get(url_prod)
    soup = bs(prod.text, 'html.parser')
    return soup


def get_url_spec(soup):
    '''возвращает url страницы по кнопке "все характеристики"
    '''
    ''' проработать принимаемый аргумент'''
    # soup = product(url_prod)
    div = soup.find('div', class_="list-more-small")
    elem = str(div).split()
    str_with_url = str(elem[6])
    list_url = str_with_url.split('=')
    urll = (list_url[1] + '=' + list_url[2]).split('"')
    url_spec = urll[1]
    return url_spec


def specifikations(url_specification):
    '''  возвращает список характеристик и их значений
    '''
    ''' принимаемый аргумент
    '''
    # url_spec = get_url_spec()
    spec = re.get(url_specification)
    soup_spec = bs(spec.content, 'html.parser')
    # t = soup_spec.find_all(text=True)
    # /// # print(t)
    spec_dict = dict()
    parameters = soup_spec.findAll('tr')
    n = -1
    key = 1
    for i in parameters:
        td = i.findAll('td')
        line = []

        for j in td:
            param = j.find_all(text=True)

            parameter = ''
            for elem in param:
                norm = ud.normalize('NFKC', elem)
                parameter += norm + ' '
            if n > 0:
                line.append(parameter)

        n += 1
        if len(line) != 0:
            if len(line[0]) != 0:
                print('line', n, key, line)
                spec_dict[key] = line
                key += 1

        '''проработать запись в эксель построчно или создание списка'''
    # for key, value in spec_dict.items():
    # print('el_dict', key, value)
    return spec_dict


# print(specifikations())


def name_product(soup):
    '''принимаемый аргумент
    '''
    # soup = product(url_prod)
    name = soup.find('h1', class_="t2 no-mobile")
    name_ = name.findAll(text=True)
    name_text = str()
    for i in name_:
        sm = ud.normalize('NFKC', i)
        name_text += sm + ' '
    # print(name_text)
    # print(name_text)
    return name_text


# print(name_product())


def only_shop(soup):
    '''
     парсит имя, url и цену в случае, если
    магазин в каталоге только один
    '''
    # soup = product(url_prod)
    block_name = soup.findAll('div', class_="wb-s-desc")
    url_name = str(block_name).split('"')
    url = url_name[15]
    name = url_name[20].split('/')[0][1:-1]
    # print('url', url)
    # print('name', name)

    price_block = soup.findAll('div', class_="desc-big-price")
    price = str(price_block).split('span')[5].split('>')[1][0:-2]

    price_only_shop = []
    list_price = []
    list_price.append(name)
    list_price.append(url)
    list_price.append(price)
    print(list_price[2])
    price_only_shop.append(list_price)
    return price_only_shop


# print(only_shop())


def get_market_url(soup):
    '''
    проработать приннимаемый аргумент,
    во избежании двойного обращения к функции product
    '''
    # soup = product(url_prod)
    url_markets = soup.findAll('table', class_="desc-hot-prices")
    price = soup.findAll('td', class_="model-shop-price")
    price_list = []
    for p in price:
        p1 = str(p).split('span')
        p2 = p1[-2]
        p2 = p2[1:-3]
        price_list.append(p2)
    all_price = []
    n_price = 0

    for shop in url_markets:
        name_shop = shop.find_all('u')
        for i in name_shop:
            name_and_url = []
            st = str(i)
            name = st[3:-4]
            a = i.parent
            url_shop = str(a).split('"')
            name_and_url.append(name)
            name_and_url.append(url_shop[9])
            all_price.append(name_and_url)
            n_price += 1

    for i in range(0, n_price):
        all_price[i].append(price_list[i])

    if len(all_price) < 1:
        all_price = only_shop(soup)

    price = []
    for i in all_price:
        shop = []
        for j in i:
            norm = ud.normalize('NFKC', j)
            shop.append(norm)
    price.append(shop)

    # print('list', all_price)  # отфармотировать значение цены
    return price


# print(get_market_url())
def open_exel(name_prod, spec_dict, market_list):
    wb = openpyxl.Workbook()
    wb.create_sheet(title='характеристики', index=0)
    sheet_spec = wb['характеристики']
    head = sheet_spec.cell(row=1, column=1)
    head.value = name_prod

    for row, key in zip(range(2, len(spec_dict) + 2), range(1, len(spec_dict) + 1)):
        for col, spec in zip(range(1, 3), range(0, len(spec_dict[key]))):
            cell = sheet_spec.cell(row=row, column=col)
            cell.value = spec_dict[key][spec]

    sheet_spec.column_dimensions['A'].width = 45

    wb.create_sheet(title='магазины', index=1)
    sheet_shop = wb['магазины']
    head = sheet_shop.cell(row=1, column=1)
    head.value = name_prod
    for row, shop in zip(range(2, len(market_list) + 2), range(0, len(market_list))):
        for col, elem in zip(range(1, 4), range(0, 3)):
            cell = sheet_shop.cell(row=row, column=col)
            cell.value = market_list[shop][elem]

    sheet_shop.column_dimensions['A'].width = 17

    wb.save(name_prod + '.xlsx')


def main():
    soup = product(url_prod)
    url_specification = get_url_spec(soup)
    spec_dict = specifikations(url_specification)
    name_prod = name_product(soup)
    market_list = get_market_url(soup)
    print('name\n', name_prod)
    print('характеристики\n', spec_dict)
    print('магазины\n', market_list)
    open_exel(name_prod, spec_dict, market_list)


if __name__ == '__main__':
    main()
